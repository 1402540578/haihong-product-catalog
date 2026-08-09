#!/usr/bin/env python3
"""Move product images into series subfolders under import/images.

Examples after organizing:
  import/images/质感针织绒/01ED-1070.webp
  import/images/科技皮革&面料/H1005.jpg

The script matches image filenames to product codes from import/products.xlsx.
It is safe to run repeatedly; images already in the correct folder are left in place.
"""
from __future__ import annotations

import csv
import hashlib
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "import"
IMAGES_DIR = IMPORT_DIR / "images"
PRODUCTS_XLSX = IMPORT_DIR / "products.xlsx"
REPORTS_DIR = ROOT / "reports"
REPORT_FILE = REPORTS_DIR / "image-organize-report.csv"
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}


def norm_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("㎡", "m2").replace("²", "2")
    return re.sub(r"[\s\-_/\\()（）\[\]【】{}：:·,，。;；]+", "", text)


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def safe_folder(value: str) -> str:
    value = clean_text(value) or "其他系列"
    value = re.sub(r'[\\/:*?"<>|]+', "-", value).strip(" .")
    return value or "其他系列"


def load_code_series() -> dict[str, str]:
    if not PRODUCTS_XLSX.exists():
        raise SystemExit(f"未找到 {PRODUCTS_XLSX.relative_to(ROOT)}")
    wb = load_workbook(PRODUCTS_XLSX, read_only=True, data_only=True)
    ws = wb["产品数据"] if "产品数据" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise SystemExit("products.xlsx 没有数据") from exc

    code_idx = None
    series_idx = None
    for idx, header in enumerate(headers):
        normalized = norm_header(header)
        if normalized in {"产品编码", "产品编号", "编码", "code"}:
            code_idx = idx
        if normalized in {"产品系列", "系列", "category"}:
            series_idx = idx
    if code_idx is None or series_idx is None:
        raise SystemExit("products.xlsx 需要包含“产品编码”和“产品系列”两列")

    mapping: dict[str, str] = {}
    for row in rows:
        if code_idx >= len(row):
            continue
        code = clean_text(row[code_idx])
        if not code:
            continue
        series = clean_text(row[series_idx]) if series_idx < len(row) else ""
        mapping[code] = safe_folder(series or "其他系列")
    wb.close()
    return mapping


def match_code(path: Path, known_codes: list[str]) -> str | None:
    stem = path.stem.strip().lower()
    for code in known_codes:
        lowered = code.lower()
        if stem == lowered:
            return code
        for delimiter in ("_", " ", "（", "(", "-主图", "-封面", "-image", "-img"):
            if stem.startswith(lowered + delimiter):
                return code
    return None


def same_file(a: Path, b: Path) -> bool:
    if not a.exists() or not b.exists() or a.stat().st_size != b.stat().st_size:
        return False
    def digest(path: Path) -> str:
        h = hashlib.sha256()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    return digest(a) == digest(b)


def unique_destination(target: Path, source: Path) -> tuple[Path, str]:
    if not target.exists():
        return target, ""
    if same_file(target, source):
        return target, "duplicate"
    for n in range(2, 1000):
        candidate = target.with_name(f"{target.stem}__{n}{target.suffix}")
        if not candidate.exists():
            return candidate, "renamed"
    raise RuntimeError(f"目标文件冲突过多：{target.name}")


def main() -> int:
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    code_series = load_code_series()
    known_codes = sorted(code_series, key=len, reverse=True)

    records: list[dict[str, str]] = []
    moved = 0
    unchanged = 0
    unmatched = 0
    duplicates = 0

    candidates = [
        p for p in IMAGES_DIR.rglob("*")
        if p.is_file() and p.name != ".gitkeep" and p.suffix.lower() in ALLOWED_EXTENSIONS
    ]

    for source in sorted(candidates, key=lambda p: str(p).lower()):
        code = match_code(source, known_codes)
        if not code:
            unmatched += 1
            records.append({
                "图片文件": str(source.relative_to(ROOT)), "产品编码": "", "产品系列": "",
                "处理结果": "未匹配", "目标位置": "", "说明": "文件名未匹配到产品编码"
            })
            continue

        series = code_series[code]
        target_dir = IMAGES_DIR / series
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / source.name

        try:
            # Already in the expected series folder.
            if source.parent.resolve() == target_dir.resolve():
                unchanged += 1
                records.append({
                    "图片文件": str(source.relative_to(ROOT)), "产品编码": code, "产品系列": series,
                    "处理结果": "无需移动", "目标位置": str(target.relative_to(ROOT)), "说明": "已在正确系列文件夹"
                })
                continue

            target, conflict = unique_destination(target, source)
            if conflict == "duplicate":
                source.unlink()
                duplicates += 1
                records.append({
                    "图片文件": str(source.relative_to(ROOT)), "产品编码": code, "产品系列": series,
                    "处理结果": "删除重复", "目标位置": str(target.relative_to(ROOT)), "说明": "目标位置已存在相同图片"
                })
                continue

            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(source), str(target))
            moved += 1
            records.append({
                "图片文件": str(source.relative_to(ROOT)), "产品编码": code, "产品系列": series,
                "处理结果": "已移动", "目标位置": str(target.relative_to(ROOT)),
                "说明": "目标重名，已自动改名" if conflict == "renamed" else ""
            })
        except Exception as exc:  # noqa: BLE001
            records.append({
                "图片文件": str(source.relative_to(ROOT)), "产品编码": code, "产品系列": series,
                "处理结果": "失败", "目标位置": str(target.relative_to(ROOT)), "说明": str(exc)
            })

    # Remove empty legacy subfolders, but retain the root and series folders.
    for folder in sorted((p for p in IMAGES_DIR.rglob("*") if p.is_dir()), key=lambda p: len(p.parts), reverse=True):
        try:
            if not any(folder.iterdir()):
                folder.rmdir()
        except OSError:
            pass

    with REPORT_FILE.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["图片文件", "产品编码", "产品系列", "处理结果", "目标位置", "说明"])
        writer.writeheader()
        writer.writerows(records)

    series_counts: defaultdict[str, int] = defaultdict(int)
    for p in IMAGES_DIR.rglob("*"):
        if p.is_file() and p.suffix.lower() in ALLOWED_EXTENSIONS:
            try:
                rel = p.relative_to(IMAGES_DIR)
                if len(rel.parts) >= 2:
                    series_counts[rel.parts[0]] += 1
            except ValueError:
                pass

    print("图片按系列整理完成")
    print(f"已移动：{moved}")
    print(f"无需移动：{unchanged}")
    print(f"删除重复：{duplicates}")
    print(f"未匹配：{unmatched}")
    print("系列统计：")
    for name, count in sorted(series_counts.items(), key=lambda x: (-x[1], x[0])):
        print(f"  {name}: {count}")
    print(f"报告：{REPORT_FILE.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
