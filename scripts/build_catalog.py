#!/usr/bin/env python3
"""Build Haihong's static GitHub Pages catalog from Excel and product images.

Inputs:
  import/products.xlsx            Main product source (first sheet or sheet named 产品数据)
  import/images/*                 Product images named with product code
  import/images.zip               Optional image archive
  import/series-status.csv        Optional series-level active/inactive rules
  import/catalog_info.json        Catalog/company metadata

Outputs:
  _site/                          Deployable static website
  reports/import-errors.csv       Non-fatal import errors and skipped rows/files
  reports/build-summary.json      Build statistics
"""
from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import re
import shutil
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("缺少 openpyxl，请运行：pip install -r scripts/requirements.txt") from exc

ROOT = Path(__file__).resolve().parents[1]
SITE_SRC = ROOT / "site"
IMPORT_DIR = ROOT / "import"
REPORT_DIR = ROOT / "reports"
OUTPUT_DIR = ROOT / "_site"
PRODUCTS_XLSX = IMPORT_DIR / "products.xlsx"
PRODUCTS_SOURCE_JSON = IMPORT_DIR / "products.source.json"
CATALOG_INFO_FILE = IMPORT_DIR / "catalog_info.json"
SERIES_STATUS_FILE = IMPORT_DIR / "series-status.csv"
IMAGES_DIR = IMPORT_DIR / "images"
IMAGES_ZIP = IMPORT_DIR / "images.zip"
PRODUCT_IMAGE_OUTPUT = OUTPUT_DIR / "assets" / "products"
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_IMAGE_BYTES = 8 * 1024 * 1024

TEXT_FIELDS = {
    "code", "salesToolType", "categoryRaw", "category", "tier", "name",
    "salesNameCn", "salesNameEn", "greyCommonName", "greyName", "backing1",
    "backing2", "alternativeBacking", "finishing", "finishingDetail",
    "qualityStandard", "greyQuality", "specialTreatment", "recommendedMarkets",
    "restrictedMarkets", "sellingPoints", "greySellingPoints", "hotColors",
    "trendColors", "moq", "mcq", "scq", "rollMeters", "developmentDate",
    "colorMethod", "projectType", "patternCode", "stockStatus", "uid",
    "imageFilename",
}
NUMBER_FIELDS = {
    "weight", "width", "greyWeight", "greyWidth", "foreignPrice",
    "domesticPrice", "textureIndex", "duplicateIndex",
}
BOOL_FIELDS = {"hasStock"}

COLUMN_ALIASES = {
    "uid": ["唯一标识", "uid"],
    "code": ["产品编码", "产品编号", "编码", "code"],
    "duplicateIndex": ["重复序号", "duplicateindex"],
    "status": ["上架状态", "产品状态", "状态", "status"],
    "salesToolType": ["销售工具", "销售工具类型", "salestooltype"],
    "categoryRaw": ["产品系列原始", "原始系列", "categoryraw"],
    "category": ["产品系列", "系列", "category"],
    "tier": ["品质定位", "品质", "tier"],
    "name": ["产品品种", "产品名称", "品种", "name"],
    "salesNameCn": ["中文销售名称", "中文名", "salesnamecn"],
    "salesNameEn": ["英文销售名称", "英文名", "salesnameen"],
    "greyCommonName": ["色坯通用名称", "坯布通用名称", "greycommonname"],
    "greyName": ["色坯名称", "坯布名称", "greyname"],
    "backing1": ["底布1", "底布/复合材料1", "底布复合材料1", "backing1"],
    "backing2": ["底布2", "底布/复合材料2", "底布复合材料2", "backing2"],
    "alternativeBacking": ["可替换底布", "替代底布", "alternativebacking"],
    "finishing": ["后整理工艺", "后整理", "finishing"],
    "finishingDetail": ["后整理说明", "后整理明细", "finishingdetail"],
    "weight": ["成品克重", "成品克重（g/㎡）", "克重", "weight"],
    "width": ["成品门幅", "成品门幅（cm）", "门幅", "width"],
    "greyWeight": ["色坯克重", "色坯克重（g/㎡）", "greyweight"],
    "greyWidth": ["色坯门幅", "色坯门幅（cm）", "greywidth"],
    "qualityStandard": ["质量标准", "qualitystandard"],
    "greyQuality": ["色坯品质", "坯布品质", "greyquality"],
    "specialTreatment": ["特殊处理", "specialtreatment"],
    "recommendedMarkets": ["建议销售市场", "建议市场", "recommendedmarkets"],
    "restrictedMarkets": ["限制销售市场", "限制市场", "restrictedmarkets"],
    "sellingPoints": ["产品卖点", "卖点", "sellingpoints"],
    "greySellingPoints": ["色坯卖点", "坯布卖点", "greysellingpoints"],
    "hotColors": ["热销颜色", "hotcolors"],
    "trendColors": ["趋势颜色", "trendcolors"],
    "moq": ["MOQ", "moq"],
    "mcq": ["MCQ", "mcq"],
    "scq": ["SCQ", "scq"],
    "rollMeters": ["每卷米数", "卷长", "rollmeters"],
    "foreignPrice": ["外贸参考价", "外贸参考价（USD/m）", "外贸价", "foreignprice"],
    "domesticPrice": ["内贸参考价", "内贸参考价（CNY/m）", "内贸价", "domesticprice"],
    "developmentDate": ["开发日期", "开发时间", "developmentdate"],
    "colorMethod": ["色彩方式", "colormethod"],
    "projectType": ["项目类型", "projecttype"],
    "patternCode": ["花型编号", "花型号", "patterncode"],
    "stockStatus": ["现货状态", "现货说明", "stockstatus"],
    "hasStock": ["有现货", "有现货/备货", "有现货 / 备货", "备货", "hasstock"],
    "imageFilename": ["图片文件名", "产品图片", "图片", "imagefilename"],
    "textureIndex": ["纹理序号", "textureindex"],
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("㎡", "m2").replace("²", "2")
    return re.sub(r"[\s\-_/\\()（）\[\]【】{}：:·,，。;；]+", "", text)


HEADER_MAP: dict[str, str] = {}
for field, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        HEADER_MAP[normalize_header(alias)] = field


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def clean_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    try:
        number = float(value)
    except (TypeError, ValueError):
        text = re.sub(r"[^0-9.\-]+", "", str(value))
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None
    return int(number) if number.is_integer() else round(number, 4)


def clean_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {
        "1", "true", "yes", "y", "是", "有", "现货", "有现货", "备货", "有备货",
    }


def clean_status(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"inactive", "0", "false", "no", "下架", "已下架", "停售", "停用"}:
        return "inactive"
    return "active"


def safe_slug(value: Any) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value or "product").strip()).strip("-._")
    return text[:80] or "product"


def stable_uid(code: str, duplicate_index: int, row_number: int) -> str:
    digest = hashlib.sha1(f"{code}|{duplicate_index}|{row_number}".encode("utf-8")).hexdigest()[:8]
    return f"{safe_slug(code)}-{duplicate_index or 1}-{digest}"


def add_error(errors: list[dict[str, str]], source: str, location: str, code: str, message: str) -> None:
    errors.append({"来源": source, "位置/文件": location, "产品编码": code, "失败原因": message})


def read_products_from_excel(errors: list[dict[str, str]]) -> list[dict[str, Any]]:
    if not PRODUCTS_XLSX.exists():
        raise FileNotFoundError(f"未找到 {PRODUCTS_XLSX.relative_to(ROOT)}")
    workbook = load_workbook(PRODUCTS_XLSX, read_only=True, data_only=True)
    sheet = workbook["产品数据"] if "产品数据" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError("products.xlsx 没有数据") from exc
    mapped_headers: list[str | None] = [HEADER_MAP.get(normalize_header(h)) for h in headers]
    if "code" not in mapped_headers:
        raise ValueError("products.xlsx 缺少“产品编码”列")

    products: list[dict[str, Any]] = []
    seen_uids: set[str] = set()
    code_occurrences: defaultdict[str, int] = defaultdict(int)
    for row_number, row in enumerate(rows, start=2):
        raw: dict[str, Any] = {}
        for index, value in enumerate(row):
            if index >= len(mapped_headers):
                break
            field = mapped_headers[index]
            if field and value not in (None, ""):
                raw[field] = value
        if not raw:
            continue
        code = clean_text(raw.get("code"))
        if not code:
            add_error(errors, "产品信息", f"第 {row_number} 行", "", "缺少产品编码，已跳过")
            continue
        code_occurrences[code] += 1
        duplicate_index = clean_number(raw.get("duplicateIndex")) or code_occurrences[code]
        duplicate_index = int(duplicate_index)
        product: dict[str, Any] = {}
        for field in TEXT_FIELDS:
            if field in raw:
                product[field] = clean_text(raw[field])
        for field in NUMBER_FIELDS:
            if field in raw:
                product[field] = clean_number(raw[field])
        for field in BOOL_FIELDS:
            if field in raw:
                product[field] = clean_bool(raw[field])
        product["code"] = code
        product["duplicateIndex"] = duplicate_index
        product["status"] = clean_status(raw.get("status"))
        product["name"] = clean_text(raw.get("name")) or clean_text(raw.get("salesNameCn")) or "未命名产品"
        product["category"] = clean_text(raw.get("category")) or "其他系列"
        product["categoryRaw"] = clean_text(raw.get("categoryRaw")) or product["category"]
        product["tier"] = clean_text(raw.get("tier")) or "待分类"
        product["hasStock"] = clean_bool(raw.get("hasStock"))
        product["textureIndex"] = int(clean_number(raw.get("textureIndex")) or (row_number % 12))
        uid = clean_text(raw.get("uid")) or stable_uid(code, duplicate_index, row_number)
        if uid in seen_uids:
            uid = stable_uid(code, duplicate_index, row_number)
        seen_uids.add(uid)
        product["uid"] = uid
        products.append(product)
    workbook.close()
    return products


def load_json_fallback(errors: list[dict[str, str]]) -> list[dict[str, Any]]:
    if not PRODUCTS_SOURCE_JSON.exists():
        return []
    try:
        data = json.loads(PRODUCTS_SOURCE_JSON.read_text("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        add_error(errors, "产品信息", str(PRODUCTS_SOURCE_JSON.name), "", f"备用 JSON 解析失败：{exc}")
        return []
    if not isinstance(data, list):
        add_error(errors, "产品信息", str(PRODUCTS_SOURCE_JSON.name), "", "备用 JSON 不是产品数组")
        return []
    return [dict(item) for item in data if isinstance(item, dict)]


def read_series_rules(errors: list[dict[str, str]]) -> dict[str, str]:
    if not SERIES_STATUS_FILE.exists():
        return {}
    rules: dict[str, str] = {}
    try:
        with SERIES_STATUS_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for line, row in enumerate(reader, start=2):
                category = clean_text(row.get("产品系列") or row.get("系列") or row.get("category"))
                action = clean_text(row.get("操作") or row.get("上架状态") or row.get("status"))
                if not category and not action:
                    continue
                if not category or not action:
                    add_error(errors, "系列上下架", f"第 {line} 行", "", "产品系列或操作为空，已跳过")
                    continue
                status = clean_status(action)
                rules[category] = status
    except (OSError, csv.Error) as exc:
        add_error(errors, "系列上下架", SERIES_STATUS_FILE.name, "", f"读取失败：{exc}")
    return rules


def iter_image_candidates(temp_dir: Path, errors: list[dict[str, str]]) -> list[Path]:
    candidates: list[Path] = []
    if IMAGES_DIR.exists():
        candidates.extend(path for path in IMAGES_DIR.rglob("*") if path.is_file() and path.name != ".gitkeep")
    if IMAGES_ZIP.exists():
        try:
            with zipfile.ZipFile(IMAGES_ZIP) as archive:
                for info in archive.infolist():
                    if info.is_dir():
                        continue
                    target = (temp_dir / info.filename).resolve()
                    try:
                        target.relative_to(temp_dir.resolve())
                    except ValueError:
                        add_error(errors, "产品图片", info.filename, "", "压缩包路径不安全，已跳过")
                        continue
                    target.parent.mkdir(parents=True, exist_ok=True)
                    target.write_bytes(archive.read(info))
            candidates.extend(path for path in temp_dir.rglob("*") if path.is_file())
        except (zipfile.BadZipFile, OSError) as exc:
            add_error(errors, "产品图片", IMAGES_ZIP.name, "", f"压缩包读取失败：{exc}")
    return candidates


def image_code_from_name(path: Path) -> str:
    stem = path.stem.strip()
    # Fallback cleanup. Numeric hyphen sections are preserved because they are often part of product codes.
    stem = re.sub(r"(?:[_\-\s（(](?:主图|封面|正面|image|img).*)$", "", stem, flags=re.IGNORECASE)
    return stem.strip()


def match_image_code(path: Path, known_codes: list[str]) -> str:
    stem = path.stem.strip().lower()
    for code in known_codes:
        lowered = code.lower()
        if stem == lowered:
            return code
        for delimiter in ("_", " ", "（", "(", "-主图", "-封面", "-image", "-img"):
            if stem.startswith(lowered + delimiter):
                return code
    return image_code_from_name(path)


def resolve_images(products: list[dict[str, Any]], errors: list[dict[str, str]]) -> None:
    PRODUCT_IMAGE_OUTPUT.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="haihong-images-") as temp:
        temp_dir = Path(temp)
        candidates = iter_image_candidates(temp_dir, errors)
        by_code: defaultdict[str, list[Path]] = defaultdict(list)
        by_name: dict[str, Path] = {}
        known_code_list = sorted({str(product.get("code") or "").strip() for product in products if product.get("code")}, key=len, reverse=True)
        for path in candidates:
            ext = path.suffix.lower()
            if ext not in ALLOWED_IMAGE_EXTENSIONS:
                add_error(errors, "产品图片", path.name, "", "不支持的图片格式，支持 JPG、PNG、WEBP")
                continue
            if path.stat().st_size > MAX_IMAGE_BYTES:
                add_error(errors, "产品图片", path.name, "", "图片超过 8MB，已跳过")
                continue
            by_name[path.name.lower()] = path
            matched_code = match_image_code(path, known_code_list)
            by_code[matched_code.lower()].append(path)

        existing_assets = SITE_SRC / "assets"
        copied_paths: dict[Path, str] = {}
        matched_files: set[Path] = set()
        for product in products:
            code = str(product.get("code") or "").strip()
            selected: Path | None = None
            explicit = clean_text(product.pop("imageFilename", None))
            if explicit:
                selected = by_name.get(Path(explicit).name.lower())
                if not selected:
                    add_error(errors, "产品图片", explicit, code, "Excel 指定的图片文件未找到")
            if not selected:
                matches = by_code.get(code.lower(), [])
                if matches:
                    selected = sorted(matches, key=lambda path: path.name.lower())[0]
            # Preserve the original bundled real image if no newly imported image exists.
            if not selected:
                for ext in ALLOWED_IMAGE_EXTENSIONS:
                    existing = existing_assets / f"{code}{ext}"
                    if existing.exists():
                        selected = existing
                        break
            if selected:
                if selected not in copied_paths:
                    ext = selected.suffix.lower()
                    destination_name = f"{safe_slug(code)}{ext}"
                    destination = PRODUCT_IMAGE_OUTPUT / destination_name
                    shutil.copy2(selected, destination)
                    copied_paths[selected] = f"assets/products/{destination_name}"
                product["image"] = copied_paths[selected]
                product["imageIsReal"] = True
                matched_files.add(selected)
            else:
                product["image"] = None
                product["imageIsReal"] = False

        known_codes = {str(product.get("code") or "").lower() for product in products}
        for path in candidates:
            if path.suffix.lower() in ALLOWED_IMAGE_EXTENSIONS and path not in matched_files:
                guessed = match_image_code(path, known_code_list)
                if guessed.lower() not in known_codes:
                    add_error(errors, "产品图片", path.name, guessed, "未匹配到产品编码")


def count_entries(products: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
    counter = Counter(str(p.get(field) or "").strip() for p in products if str(p.get(field) or "").strip())
    return [{"name": name, "count": count} for name, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))]


def build_catalog_payload(products: list[dict[str, Any]]) -> dict[str, Any]:
    active = [product for product in products if product.get("status") != "inactive"]
    categories = count_entries(active, "category")
    tiers = count_entries(active, "tier")
    sales_types = count_entries(active, "salesToolType")
    names = {str(product.get("name") or "").strip() for product in active if product.get("name")}
    prices = [float(product["foreignPrice"]) for product in active if isinstance(product.get("foreignPrice"), (int, float))]
    try:
        catalog_info = json.loads(CATALOG_INFO_FILE.read_text("utf-8")) if CATALOG_INFO_FILE.exists() else {}
    except (json.JSONDecodeError, UnicodeDecodeError):
        catalog_info = {}
    now = dt.datetime.now(dt.timezone.utc)
    meta = {
        **catalog_info,
        "sourceFile": "import/products.xlsx",
        "sourceDate": now.date().isoformat(),
        "generatedAt": now.isoformat(timespec="seconds"),
        "productCount": len(active),
        "allProductCount": len(products),
        "inactiveProductCount": len(products) - len(active),
        "categoryCount": len(categories),
        "productNameCount": len(names),
        "priceMin": min(prices) if prices else None,
        "priceMax": max(prices) if prices else None,
        "categories": categories,
        "tiers": tiers,
        "salesToolTypes": sales_types,
    }
    public_products = []
    for product in active:
        clean = {key: value for key, value in product.items() if key not in {"status", "createdAt", "updatedAt"}}
        public_products.append(clean)
    return {"meta": meta, "products": public_products}


def write_reports(errors: list[dict[str, str]], summary: dict[str, Any]) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    error_file = REPORT_DIR / "import-errors.csv"
    with error_file.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["来源", "位置/文件", "产品编码", "失败原因"])
        writer.writeheader()
        writer.writerows(errors)
    (REPORT_DIR / "build-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), "utf-8"
    )


def prepare_output() -> None:
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    shutil.copytree(SITE_SRC, OUTPUT_DIR)
    (OUTPUT_DIR / "data").mkdir(parents=True, exist_ok=True)
    PRODUCT_IMAGE_OUTPUT.mkdir(parents=True, exist_ok=True)
    # GitHub Pages should not run Jekyll processing.
    (OUTPUT_DIR / ".nojekyll").write_text("", "utf-8")
    shutil.copy2(OUTPUT_DIR / "index.html", OUTPUT_DIR / "404.html")


def main() -> int:
    errors: list[dict[str, str]] = []
    prepare_output()
    try:
        products = read_products_from_excel(errors)
    except (FileNotFoundError, ValueError) as exc:
        products = load_json_fallback(errors)
        if not products:
            print(f"构建失败：{exc}", file=sys.stderr)
            write_reports(errors, {"ok": False, "fatalError": str(exc)})
            return 1
        add_error(errors, "产品信息", "products.xlsx", "", f"Excel 无法使用，已采用备用 JSON：{exc}")

    rules = read_series_rules(errors)
    affected_by_rules = 0
    if rules:
        for product in products:
            category = str(product.get("category") or "").strip()
            if category in rules:
                product["status"] = rules[category]
                affected_by_rules += 1

    resolve_images(products, errors)
    payload = build_catalog_payload(products)
    (OUTPUT_DIR / "data" / "catalog.json").write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")), "utf-8"
    )
    summary = {
        "ok": True,
        "generatedAt": payload["meta"]["generatedAt"],
        "totalRows": len(products),
        "publishedProducts": payload["meta"]["productCount"],
        "inactiveProducts": payload["meta"]["inactiveProductCount"],
        "productsWithImages": sum(1 for product in payload["products"] if product.get("image")),
        "seriesRules": len(rules),
        "productsAffectedBySeriesRules": affected_by_rules,
        "warnings": len(errors),
        "siteDirectory": "_site",
    }
    write_reports(errors, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if errors:
        print(f"构建完成，但有 {len(errors)} 条提示；请下载 reports/import-errors.csv 查看。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
